from datetime import datetime, timedelta
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io
from app.models import User, Meal, GlucoseReading, InsulinDose

async def generate_diary_text(session: AsyncSession, user: User, days: int = 7) -> str:
    """Generate text view of diary for N days"""
    now = datetime.now()
    start_date = now - timedelta(days=days)

    # Get all data
    meals_result = await session.execute(
        select(Meal).where(
            Meal.user_id == user.id,
            Meal.datetime >= start_date
        ).order_by(Meal.datetime)
    )
    meals = meals_result.scalars().all()

    glucose_result = await session.execute(
        select(GlucoseReading).where(
            GlucoseReading.user_id == user.id,
            GlucoseReading.datetime >= start_date
        ).order_by(GlucoseReading.datetime)
    )
    glucose_readings = glucose_result.scalars().all()

    insulin_result = await session.execute(
        select(InsulinDose).where(
            InsulinDose.user_id == user.id,
            InsulinDose.datetime >= start_date
        ).order_by(InsulinDose.datetime)
    )
    insulin_doses = insulin_result.scalars().all()

    if not meals and not glucose_readings and not insulin_doses:
        return "📊 Дневник пуст за указанный период."

    # Build text diary
    text = f"📊 **Дневник за {days} дн.**\n"
    text += f"Период: {start_date.strftime('%d.%m')} — {now.strftime('%d.%m.%Y')}\n\n"

    # Group by date
    dates = {}

    for meal in meals:
        date_key = meal.datetime.date()
        if date_key not in dates:
            dates[date_key] = {'meals': [], 'glucose': [], 'insulin': []}
        dates[date_key]['meals'].append(meal)

    for reading in glucose_readings:
        date_key = reading.datetime.date()
        if date_key not in dates:
            dates[date_key] = {'meals': [], 'glucose': [], 'insulin': []}
        dates[date_key]['glucose'].append(reading)

    for dose in insulin_doses:
        date_key = dose.datetime.date()
        if date_key not in dates:
            dates[date_key] = {'meals': [], 'glucose': [], 'insulin': []}
        dates[date_key]['insulin'].append(dose)

    for date_key in sorted(dates.keys(), reverse=True):
        data = dates[date_key]
        text += f"**{date_key.strftime('%d.%m.%Y')}**\n"

        # Fasting glucose first
        fasting = [r for r in data['glucose'] if r.type.value == 'fasting']
        fasting.sort(key=lambda r: r.datetime)
        if fasting:
            r = fasting[0]
            status = "✅" if r.is_normal else "⚠️"
            text += f"  🩸 {r.datetime.strftime('%H:%M')} Натощак: {r.value} {status}\n"
        else:
            text += f"  🩸 Натощак: —\n"

        # Meals sorted by time, each as a block with linked glucose and insulin
        day_meals = sorted(data['meals'], key=lambda m: m.datetime)
        postmeal_readings = [r for r in data['glucose'] if r.type.value == 'postmeal']
        insulin_doses_day = sorted(data['insulin'], key=lambda d: d.datetime)

        # Track which readings/doses are used
        used_readings = set()
        used_doses = set()

        for meal in day_meals:
            text += f"\n  🍽 {meal.datetime.strftime('%H:%M')} {meal.description}\n"

            # Find linked glucose (by meal_id or closest postmeal within 3h after meal)
            linked_reading = None
            for r in postmeal_readings:
                if r.id in used_readings:
                    continue
                if r.meal_id == meal.id:
                    linked_reading = r
                    break
            if not linked_reading:
                for r in postmeal_readings:
                    if r.id in used_readings:
                        continue
                    delta = (r.datetime - meal.datetime).total_seconds()
                    if 0 < delta <= 10800:
                        linked_reading = r
                        break

            if linked_reading:
                used_readings.add(linked_reading.id)
                status = "✅" if linked_reading.is_normal else "⚠️"
                text += f"  🔬 {linked_reading.datetime.strftime('%H:%M')} Сахар: {linked_reading.value} {status}\n"
            else:
                text += f"  🔬 Сахар: —\n"

            # Find closest insulin within 1h before to 2h after meal
            linked_dose = None
            for d in insulin_doses_day:
                if d.id in used_doses:
                    continue
                delta = (d.datetime - meal.datetime).total_seconds()
                if -3600 <= delta <= 7200:
                    linked_dose = d
                    break

            if linked_dose:
                used_doses.add(linked_dose.id)
                type_text = "кор." if linked_dose.type.value == "short" else "длин."
                text += f"  💉 {linked_dose.datetime.strftime('%H:%M')} Инсулин: {linked_dose.units} Ед ({type_text})\n"
            else:
                text += f"  💉 Инсулин: —\n"

        # Orphan insulin/glucose not linked to any meal
        orphan_readings = [r for r in postmeal_readings if r.id not in used_readings]
        orphan_doses = [d for d in insulin_doses_day if d.id not in used_doses]
        if orphan_readings or orphan_doses:
            for r in orphan_readings:
                status = "✅" if r.is_normal else "⚠️"
                text += f"\n  🔬 {r.datetime.strftime('%H:%M')} Сахар: {r.value} {status}\n"
            for d in orphan_doses:
                type_text = "кор." if d.type.value == "short" else "длин."
                text += f"  💉 {d.datetime.strftime('%H:%M')} Инсулин: {d.units} Ед ({type_text})\n"

        text += "\n"

    return text

async def generate_excel_export(session: AsyncSession, user: User, days: int = 7) -> io.BytesIO:
    """Generate Excel export for doctor"""
    now = datetime.now()
    start_date = now - timedelta(days=days)

    # Get all data
    meals_result = await session.execute(
        select(Meal).where(
            Meal.user_id == user.id,
            Meal.datetime >= start_date
        ).order_by(Meal.datetime)
    )
    meals = {m.id: m for m in meals_result.scalars().all()}

    glucose_result = await session.execute(
        select(GlucoseReading).where(
            GlucoseReading.user_id == user.id,
            GlucoseReading.datetime >= start_date
        ).order_by(GlucoseReading.datetime)
    )
    glucose_readings = glucose_result.scalars().all()

    insulin_result = await session.execute(
        select(InsulinDose).where(
            InsulinDose.user_id == user.id,
            InsulinDose.datetime >= start_date
        ).order_by(InsulinDose.datetime)
    )
    insulin_doses = insulin_result.scalars().all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Дневник ГСД"

    # Header
    header_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws['A1'] = 'ДНЕВНИК ГСД'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Пациент: {user.full_name or user.username or "Пользователь"}'
    ws['A3'] = f'Период: {start_date.strftime("%d.%m.%Y")} — {now.strftime("%d.%m.%Y")}'

    # Table headers
    ws['A5'] = 'Дата'
    ws['B5'] = 'Время'
    ws['C5'] = 'Приём пищи'
    ws['D5'] = 'Сахар'
    ws['E5'] = 'Инсулин'

    for cell in ['A5', 'B5', 'C5', 'D5', 'E5']:
        ws[cell].font = header_font
        ws[cell].border = border
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # Collect all events
    events = []

    for meal in meals.values():
        events.append({
            'datetime': meal.datetime,
            'type': 'meal',
            'data': meal
        })

    for reading in glucose_readings:
        events.append({
            'datetime': reading.datetime,
            'type': 'glucose',
            'data': reading,
            'meal_id': reading.meal_id
        })

    for dose in insulin_doses:
        events.append({
            'datetime': dose.datetime,
            'type': 'insulin',
            'data': dose
        })

    def excel_sort_key(event):
        is_fasting = (event['type'] == 'glucose' and event['data'].type.value == 'fasting')
        return (event['datetime'].date(), 0 if is_fasting else 1, event['datetime'])

    events.sort(key=excel_sort_key)

    # Fill rows
    row = 6
    current_date = None

    for event in events:
        dt = event['datetime']
        date_str = dt.strftime('%d.%m.%Y')
        time_str = dt.strftime('%H:%M')

        # Date only on first row of the day
        if current_date != dt.date():
            ws[f'A{row}'] = date_str
            current_date = dt.date()

        ws[f'B{row}'] = time_str

        if event['type'] == 'meal':
            ws[f'C{row}'] = event['data'].description
        elif event['type'] == 'glucose':
            status = "✅" if event['data'].is_normal else "⚠️"
            ws[f'D{row}'] = f"{event['data'].value} {status}"
        elif event['type'] == 'insulin':
            type_text = "К" if event['data'].type.value == "short" else "Д"
            ws[f'E{row}'] = f"{event['data'].units} Ед ({type_text})"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].border = border

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output
